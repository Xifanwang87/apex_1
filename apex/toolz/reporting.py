# Creation date:  5/15/2017
# Author:         Eduardo Sahione
# Author's email: esahione@salientpartners.com
# Copyright Salient Partners, LP
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import *
import pandas as pd


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None, number_format=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)


    rows = ws[cell_range]
    for row in rows:
        for c in row:
            if font:
                c.font = font
            if alignment:
                c.alignment = alignment
            if number_format:
                c.number_format = number_format

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def as_text(value):
    return str(value) if value is not None else ""

class SalientReportWorkbook(object):
    """
    This class helps hold the data for a particular sheet in an excel file.
    """
    def __init__(self):
        self._wb = Workbook()
        self._wb.remove(self._wb.active)
        self._sheet_makers = []

    def save(self, filename):
        for maker in self._sheet_makers:
            for column_cells in maker.worksheet.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                maker.worksheet.column_dimensions[column_cells[0].column].width = length *  1.4 # Due to font size
        self._wb.save(filename)

    def add_sheet(self, sheet_maker):
        self._sheet_makers.append(sheet_maker)
        sheet_maker.initialize(self._wb)
        sheet_maker.write()

class ReportWorksheet(object):
    """
    This class helps hold the data for a particular sheet in an excel file.

    """
    def __init__(self, title, index=True):
        self.title = title
        self.cursor = (0, 0)
        self.worksheet = None
        self.index = index

    def initialize(self, wb):
        self.worksheet = wb.create_sheet(title=self.title)

    def write(self):
        raise NotImplementedError

    def create_styles(self):
        pass

    def create_chart_style(self):
        pass


class SalientDataframeExcelReport(ReportWorksheet):
    def __init__(self, data, title, emphasis_index_name=None, emphasis_column_name=None, percentage_cols=[], index=True):
        super().__init__(title, index=index)
        self.data = data
        self.emphasis_index_name = emphasis_index_name
        self.emphasis_column_name = emphasis_column_name
        if not isinstance(percentage_cols, list):
            self.percentage_cols = [percentage_cols]
        self.percentage_cols = percentage_cols

    def write(self):
        rc = 0
        for r in dataframe_to_rows(self.data, index=self.index, header=True):
            rc += 1
            if rc == 2:
                continue
            self.worksheet.append(r)

        white_fill = PatternFill("solid", fgColor="FFFFFF")
        header_font = Font(name='Rockwell', size=11, bold=True, color="FFFFFF")
        header_emphasis_font = Font(name='Rockwell', size=12, bold=True, color="000000")
        header_emphasis_fill = PatternFill("solid", fgColor="25BBB8")
        header_non_emphasis_fill = PatternFill("solid", fgColor="3C2E48")
        header_border = Side(style='thin', color="000000")
        text_font = Font(name='Arial', size=9, bold=False, color="898989")
        text_emphasis_font = Font(name='Arial', size=10, bold=True, color="898989")
        index_font = Font(name='Rockwell', size=11, bold=True, color="000000")
        index_emphasis_font = Font(name='Rockwell', size=12, bold=True, color="000000")
        index_emphasis_fill = PatternFill("solid", fgColor="25BBB8")
        index_non_emphasis_fill = PatternFill("solid", fgColor="FFFFFF")
        index_border = Side(style='thin', color="000000")
        last_col = len(self.data.columns)
        last_row = len(self.data) + 1
        if self.index:
            last_col += 1
        if self.index:
            header_range = self.worksheet.iter_cols(min_col=2, max_col=last_col, max_row=1, min_row=1)
        else:
            header_range = self.worksheet.iter_cols(min_col=1, max_col=last_col, max_row=1, min_row=1)
        index_range = self.worksheet.iter_rows(min_col=1, max_col=1, max_row=last_row, min_row=2)


        # First let's style the actual table data
        if self.index:
            style_range(self.worksheet, f'B2:{get_column_letter(last_col)}{last_row}',
                font=text_font,
                alignment=Alignment(horizontal='center'),
                fill=white_fill,
                number_format='#,##0.00'
            )
        else:
            style_range(self.worksheet, f'A2:{get_column_letter(last_col)}{last_row}',
                font=text_font,
                alignment=Alignment(horizontal='center'),
                fill=white_fill,
                number_format='#,##0.00'
            )
        # Ṅow let's style the header
        style_range(self.worksheet, f'A1:{get_column_letter(last_col)}1',
            font=header_font,
            alignment=Alignment(horizontal='center'),
            fill=header_non_emphasis_fill,
            border=Border(bottom=header_border)
        )
        if self.index:
            # And the index
            if isinstance(self.data.index, pd.DatetimeIndex):
                index_number_format = 'M-DD-YYYY'
            else:
                index_number_format = '#,##0.00'

            style_range(self.worksheet, f'A1:A{last_row}',
                font=index_font,
                alignment=Alignment(horizontal='center'),
                fill=index_non_emphasis_fill,
                border=Border(right=header_border),
                number_format=index_number_format
            )
        for col in self.percentage_cols:
            col_num = list(self.data.columns).index(col) + 1
            if self.index:
                col_num += 1
            col_letter = get_column_letter(col_num)
            style_range(self.worksheet, f'{col_letter}2:{col_letter}{last_row}',
                font=text_font,
                alignment=Alignment(horizontal='center'),
                fill=white_fill,
                number_format='0.00%'
            )
        if self.emphasis_column_name is not None:
            try:
                emphasis_col = list(self.data.columns).index(self.emphasis_column_name) + 2
                header_cell = self.worksheet[f'{get_column_letter(emphasis_col)}1']
                header_cell.font = header_emphasis_font
                header_cell.fill = header_emphasis_fill
                style_range(self.worksheet, f'{get_column_letter(emphasis_col)}2:{get_column_letter(emphasis_col)}{last_row}',
                    font=text_emphasis_font,
                    border=Border(bottom=header_border, top=header_border, right=header_border)
                    )
            except:
                pass
        if self.emphasis_index_name is not None and self.index:
            try:
                emphasis_row = list(self.data.index).index(self.emphasis_index_name) + 2
                index_cell = self.worksheet[f'A{emphasis_row}']
                index_cell.font = index_emphasis_font
                index_cell.fill = index_emphasis_fill
                index_cell.border = Border(bottom=header_border, top=header_border, right=header_border)
                style_range(self.worksheet, f'B{emphasis_row}:{get_column_letter(last_col)}{emphasis_row}',
                    font=text_emphasis_font,
                    border=Border(bottom=header_border, top=header_border)
                    )
            except:
                pass
        ### Now setting borders throughout the range.
        style_range(self.worksheet, f'A1:{get_column_letter(last_col)}{last_row}',
            border=Border(bottom=header_border, top=header_border, left=header_border, right=header_border)
            )




class SalientExcelLineChartSheet(ReportWorksheet):
    def __init__(self, data, title):
        super().__init__(title)
        self.data = data

    def initialize(self, wb):
        self.chartsheet = wb.create_chartsheet(title=self.title)
        self.worksheet = wb.create_sheet(title=self.title + ' data')


    def write(self):
        chart = LineChart()
        for r in dataframe_to_rows(self.data, index=True, header=True):
            self.worksheet.append(r)
        data = Reference(self.worksheet, min_col=2, min_row=1, max_col=len(self.data.columns), max_row=len(self.data))
        c1 = LineChart()
        c1.title = ""
        c1.style = 13
        c1.y_axis.title = self.title
        c1.x_axis.title = 'Date'
        c1.add_data(data, titles_from_data=True)
        self.chartsheet.add_chart(c1)
